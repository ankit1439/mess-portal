from flask import Blueprint, jsonify, request
from functools import wraps
from src.models.mess_models import (
    db, AdminUser, AdminSession, Vote, Feedback, Complaint, MenuSuggestion,
    cleanup_expired_sessions
)
from datetime import datetime, timedelta
import secrets
import jwt
import os

admin_bp = Blueprint('admin', __name__)

# JWT Configuration
JWT_SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'mess_portal_jwt_secret_2024')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 1

def get_client_ip():
    """Get client IP address from request"""
    if request.environ.get('HTTP_X_FORWARDED_FOR') is None:
        return request.environ['REMOTE_ADDR']
    else:
        return request.environ['HTTP_X_FORWARDED_FOR']

def generate_jwt_token(admin_id, expires_in_hours=JWT_EXPIRATION_HOURS):
    """Generate JWT token for admin authentication"""
    payload = {
        'admin_id': admin_id,
        'exp': datetime.utcnow() + timedelta(hours=expires_in_hours),
        'iat': datetime.utcnow()
    }
    return jwt.encode(payload, JWT_SECRET_KEY, algorithm=JWT_ALGORITHM)

def verify_jwt_token(token):
    """Verify JWT token and return admin_id"""
    try:
        payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=[JWT_ALGORITHM])
        return payload['admin_id']
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None

def require_admin_auth(f):
    """Decorator to require admin authentication"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Clean up expired sessions first
        cleanup_expired_sessions()
        
        # Get token from Authorization header
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({
                'success': False,
                'error': {
                    'code': 'UNAUTHORIZED',
                    'message': 'Missing or invalid authorization header'
                }
            }), 401
        
        token = auth_header.split(' ')[1]
        
        # Verify JWT token
        admin_id = verify_jwt_token(token)
        if not admin_id:
            return jsonify({
                'success': False,
                'error': {
                    'code': 'INVALID_TOKEN',
                    'message': 'Invalid or expired token'
                }
            }), 401
        
        # Check if admin exists and is active
        admin = AdminUser.query.get(admin_id)
        if not admin or not admin.is_active:
            return jsonify({
                'success': False,
                'error': {
                    'code': 'ADMIN_NOT_FOUND',
                    'message': 'Admin user not found or inactive'
                }
            }), 401
        
        # Add admin to request context
        request.current_admin = admin
        
        return f(*args, **kwargs)
    
    return decorated_function

@admin_bp.route('/login', methods=['POST'])
def admin_login():
    """Admin login endpoint"""
    try:
        data = request.json
        
        # Validate required fields
        if not data or 'username' not in data or 'password' not in data:
            return jsonify({
                'success': False,
                'error': {
                    'code': 'MISSING_CREDENTIALS',
                    'message': 'Username and password are required'
                }
            }), 400
        
        # Find admin user
        admin = AdminUser.query.filter_by(username=data['username']).first()
        
        if not admin or not admin.verify_password(data['password']):
            return jsonify({
                'success': False,
                'error': {
                    'code': 'INVALID_CREDENTIALS',
                    'message': 'Invalid username or password'
                }
            }), 401
        
        if not admin.is_active:
            return jsonify({
                'success': False,
                'error': {
                    'code': 'ACCOUNT_DISABLED',
                    'message': 'Admin account is disabled'
                }
            }), 401
        
        # Generate JWT token
        token = generate_jwt_token(admin.id)
        
        # Update last login
        admin.last_login = datetime.utcnow()
        
        # Create session record
        ip_address = get_client_ip()
        session = AdminSession.create_session(admin.id, ip_address)
        session.token = token
        
        db.session.add(session)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'token': token,
            'expires_in': JWT_EXPIRATION_HOURS * 3600,  # in seconds
            'admin': {
                'id': admin.id,
                'username': admin.username,
                'email': admin.email
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': {
                'code': 'INTERNAL_ERROR',
                'message': 'An error occurred during login'
            }
        }), 500

@admin_bp.route('/logout', methods=['POST'])
@require_admin_auth
def admin_logout():
    """Admin logout endpoint"""
    try:
        # Get token from Authorization header
        auth_header = request.headers.get('Authorization')
        token = auth_header.split(' ')[1]
        
        # Find and delete session
        session = AdminSession.query.filter_by(token=token).first()
        if session:
            db.session.delete(session)
            db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Logged out successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': {
                'code': 'INTERNAL_ERROR',
                'message': 'An error occurred during logout'
            }
        }), 500

@admin_bp.route('/profile', methods=['GET'])
@require_admin_auth
def get_admin_profile():
    """Get current admin profile"""
    try:
        admin = request.current_admin
        return jsonify({
            'success': True,
            'admin': admin.to_dict()
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': {
                'code': 'INTERNAL_ERROR',
                'message': 'An error occurred while fetching profile'
            }
        }), 500

@admin_bp.route('/change-password', methods=['POST'])
@require_admin_auth
def change_password():
    """Change admin password"""
    try:
        data = request.json
        admin = request.current_admin
        
        # Validate required fields
        required_fields = ['current_password', 'new_password']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({
                    'success': False,
                    'error': {
                        'code': 'MISSING_FIELD',
                        'message': f'Missing required field: {field}'
                    }
                }), 400
        
        # Verify current password
        if not admin.verify_password(data['current_password']):
            return jsonify({
                'success': False,
                'error': {
                    'code': 'INVALID_PASSWORD',
                    'message': 'Current password is incorrect'
                }
            }), 400
        
        # Validate new password
        if len(data['new_password']) < 6:
            return jsonify({
                'success': False,
                'error': {
                    'code': 'WEAK_PASSWORD',
                    'message': 'New password must be at least 6 characters long'
                }
            }), 400
        
        # Update password
        admin.password_hash = AdminUser.hash_password(data['new_password'])
        db.session.commit()
        
        # Invalidate all existing sessions for this admin
        AdminSession.query.filter_by(admin_id=admin.id).delete()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Password changed successfully. Please login again.'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': {
                'code': 'INTERNAL_ERROR',
                'message': 'An error occurred while changing password'
            }
        }), 500

@admin_bp.route('/verify-token', methods=['POST'])
def verify_token():
    """Verify if a token is valid"""
    try:
        # Get token from Authorization header
        auth_header = request.headers.get('Authorization')
        if not auth_header or not auth_header.startswith('Bearer '):
            return jsonify({
                'success': False,
                'valid': False,
                'message': 'Missing or invalid authorization header'
            }), 200
        
        token = auth_header.split(' ')[1]
        
        # Verify JWT token
        admin_id = verify_jwt_token(token)
        if not admin_id:
            return jsonify({
                'success': True,
                'valid': False,
                'message': 'Invalid or expired token'
            }), 200
        
        # Check if admin exists and is active
        admin = AdminUser.query.get(admin_id)
        if not admin or not admin.is_active:
            return jsonify({
                'success': True,
                'valid': False,
                'message': 'Admin user not found or inactive'
            }), 200
        
        return jsonify({
            'success': True,
            'valid': True,
            'admin': {
                'id': admin.id,
                'username': admin.username,
                'email': admin.email
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': {
                'code': 'INTERNAL_ERROR',
                'message': 'An error occurred while verifying token'
            }
        }), 500

# Error handlers for admin routes
@admin_bp.errorhandler(401)
def unauthorized(error):
    return jsonify({
        'success': False,
        'error': {
            'code': 'UNAUTHORIZED',
            'message': 'Authentication required'
        }
    }), 401

@admin_bp.errorhandler(403)
def forbidden(error):
    return jsonify({
        'success': False,
        'error': {
            'code': 'FORBIDDEN',
            'message': 'Access denied'
        }
    }), 403



# Admin data viewing endpoints

@admin_bp.route('/votes', methods=['GET'])
@require_admin_auth
def get_all_votes():
    """Get all votes with filtering and pagination"""
    try:
        # Get query parameters
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 50, type=int)
        day = request.args.get('day')
        meal = request.args.get('meal')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Build query
        query = Vote.query
        
        # Apply filters
        if day:
            query = query.filter(Vote.day == day.lower())
        if meal:
            query = query.filter(Vote.meal == meal.lower())
        if start_date:
            try:
                start_date_obj = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                query = query.filter(Vote.timestamp >= start_date_obj)
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': {
                        'code': 'INVALID_DATE',
                        'message': 'Invalid start_date format. Use ISO format.'
                    }
                }), 400
        if end_date:
            try:
                end_date_obj = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                query = query.filter(Vote.timestamp <= end_date_obj)
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': {
                        'code': 'INVALID_DATE',
                        'message': 'Invalid end_date format. Use ISO format.'
                    }
                }), 400
        
        # Order by timestamp (newest first)
        query = query.order_by(Vote.timestamp.desc())
        
        # Get total count
        total = query.count()
        
        # Apply pagination
        votes = query.paginate(
            page=page, 
            per_page=limit, 
            error_out=False
        ).items
        
        # Calculate pagination info
        total_pages = (total + limit - 1) // limit
        
        return jsonify({
            'success': True,
            'data': [vote.to_dict() for vote in votes],
            'pagination': {
                'page': page,
                'limit': limit,
                'total': total,
                'pages': total_pages,
                'has_next': page < total_pages,
                'has_prev': page > 1
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': {
                'code': 'INTERNAL_ERROR',
                'message': 'An error occurred while fetching votes'
            }
        }), 500

@admin_bp.route('/feedback', methods=['GET'])
@require_admin_auth
def get_all_feedback():
    """Get all feedback with filtering and pagination"""
    try:
        # Get query parameters
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 50, type=int)
        feedback_type = request.args.get('feedback_type')
        rating = request.args.get('rating', type=int)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Build query
        query = Feedback.query
        
        # Apply filters
        if feedback_type:
            query = query.filter(Feedback.feedback_type == feedback_type)
        if rating:
            query = query.filter(Feedback.rating == rating)
        if start_date:
            try:
                start_date_obj = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                query = query.filter(Feedback.timestamp >= start_date_obj)
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': {
                        'code': 'INVALID_DATE',
                        'message': 'Invalid start_date format. Use ISO format.'
                    }
                }), 400
        if end_date:
            try:
                end_date_obj = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                query = query.filter(Feedback.timestamp <= end_date_obj)
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': {
                        'code': 'INVALID_DATE',
                        'message': 'Invalid end_date format. Use ISO format.'
                    }
                }), 400
        
        # Order by timestamp (newest first)
        query = query.order_by(Feedback.timestamp.desc())
        
        # Get total count
        total = query.count()
        
        # Apply pagination
        feedback_items = query.paginate(
            page=page, 
            per_page=limit, 
            error_out=False
        ).items
        
        # Calculate pagination info
        total_pages = (total + limit - 1) // limit
        
        return jsonify({
            'success': True,
            'data': [feedback.to_dict() for feedback in feedback_items],
            'pagination': {
                'page': page,
                'limit': limit,
                'total': total,
                'pages': total_pages,
                'has_next': page < total_pages,
                'has_prev': page > 1
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': {
                'code': 'INTERNAL_ERROR',
                'message': 'An error occurred while fetching feedback'
            }
        }), 500

@admin_bp.route('/complaints', methods=['GET'])
@require_admin_auth
def get_all_complaints():
    """Get all complaints with filtering and pagination"""
    try:
        # Get query parameters
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 50, type=int)
        category = request.args.get('category')
        urgency = request.args.get('urgency')
        status = request.args.get('status')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Build query
        query = Complaint.query
        
        # Apply filters
        if category:
            query = query.filter(Complaint.category == category)
        if urgency:
            query = query.filter(Complaint.urgency == urgency.lower())
        if status:
            query = query.filter(Complaint.status == status.lower())
        if start_date:
            try:
                start_date_obj = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                query = query.filter(Complaint.timestamp >= start_date_obj)
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': {
                        'code': 'INVALID_DATE',
                        'message': 'Invalid start_date format. Use ISO format.'
                    }
                }), 400
        if end_date:
            try:
                end_date_obj = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                query = query.filter(Complaint.timestamp <= end_date_obj)
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': {
                        'code': 'INVALID_DATE',
                        'message': 'Invalid end_date format. Use ISO format.'
                    }
                }), 400
        
        # Order by urgency (urgent first) then by timestamp (newest first)
        urgency_order = db.case(
            (Complaint.urgency == 'urgent', 1),
            (Complaint.urgency == 'high', 2),
            (Complaint.urgency == 'medium', 3),
            (Complaint.urgency == 'low', 4),
            else_=5
        )
        query = query.order_by(urgency_order, Complaint.timestamp.desc())
        
        # Get total count
        total = query.count()
        
        # Apply pagination
        complaints = query.paginate(
            page=page, 
            per_page=limit, 
            error_out=False
        ).items
        
        # Calculate pagination info
        total_pages = (total + limit - 1) // limit
        
        return jsonify({
            'success': True,
            'data': [complaint.to_dict() for complaint in complaints],
            'pagination': {
                'page': page,
                'limit': limit,
                'total': total,
                'pages': total_pages,
                'has_next': page < total_pages,
                'has_prev': page > 1
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': {
                'code': 'INTERNAL_ERROR',
                'message': 'An error occurred while fetching complaints'
            }
        }), 500

@admin_bp.route('/menu-suggestions', methods=['GET'])
@require_admin_auth
def get_all_menu_suggestions():
    """Get all menu suggestions with filtering and pagination"""
    try:
        # Get query parameters
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 50, type=int)
        meal_type = request.args.get('meal_type')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Build query
        query = MenuSuggestion.query
        
        # Apply filters
        if meal_type:
            query = query.filter(MenuSuggestion.meal_type == meal_type.lower())
        if start_date:
            try:
                start_date_obj = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                query = query.filter(MenuSuggestion.timestamp >= start_date_obj)
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': {
                        'code': 'INVALID_DATE',
                        'message': 'Invalid start_date format. Use ISO format.'
                    }
                }), 400
        if end_date:
            try:
                end_date_obj = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                query = query.filter(MenuSuggestion.timestamp <= end_date_obj)
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': {
                        'code': 'INVALID_DATE',
                        'message': 'Invalid end_date format. Use ISO format.'
                    }
                }), 400
        
        # Order by timestamp (newest first)
        query = query.order_by(MenuSuggestion.timestamp.desc())
        
        # Get total count
        total = query.count()
        
        # Apply pagination
        suggestions = query.paginate(
            page=page, 
            per_page=limit, 
            error_out=False
        ).items
        
        # Calculate pagination info
        total_pages = (total + limit - 1) // limit
        
        return jsonify({
            'success': True,
            'data': [suggestion.to_dict() for suggestion in suggestions],
            'pagination': {
                'page': page,
                'limit': limit,
                'total': total,
                'pages': total_pages,
                'has_next': page < total_pages,
                'has_prev': page > 1
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': {
                'code': 'INTERNAL_ERROR',
                'message': 'An error occurred while fetching menu suggestions'
            }
        }), 500

@admin_bp.route('/dashboard', methods=['GET'])
@require_admin_auth
def get_dashboard_stats():
    """Get dashboard statistics"""
    try:
        # Get current date for today's stats
        today = datetime.utcnow().date()
        
        # Basic counts
        total_votes = Vote.query.count()
        total_feedback = Feedback.query.count()
        total_complaints = Complaint.query.count()
        total_suggestions = MenuSuggestion.query.count()
        
        # Today's votes
        votes_today = Vote.query.filter(
            db.func.date(Vote.timestamp) == today
        ).count()
        
        # Popular dishes (top 10)
        popular_dishes = db.session.query(
            Vote.dish,
            db.func.count(Vote.dish).label('vote_count')
        ).group_by(Vote.dish).order_by(
            db.func.count(Vote.dish).desc()
        ).limit(10).all()
        
        # Feedback ratings distribution
        rating_distribution = db.session.query(
            Feedback.rating,
            db.func.count(Feedback.rating).label('count')
        ).filter(
            Feedback.rating.isnot(None)
        ).group_by(Feedback.rating).order_by(Feedback.rating).all()
        
        # Complaints by urgency
        complaints_by_urgency = db.session.query(
            Complaint.urgency,
            db.func.count(Complaint.urgency).label('count')
        ).group_by(Complaint.urgency).all()
        
        # Recent activity (last 7 days)
        seven_days_ago = datetime.utcnow() - timedelta(days=7)
        recent_votes = Vote.query.filter(Vote.timestamp >= seven_days_ago).count()
        recent_feedback = Feedback.query.filter(Feedback.timestamp >= seven_days_ago).count()
        recent_complaints = Complaint.query.filter(Complaint.timestamp >= seven_days_ago).count()
        
        return jsonify({
            'success': True,
            'stats': {
                'totals': {
                    'votes': total_votes,
                    'feedback': total_feedback,
                    'complaints': total_complaints,
                    'menu_suggestions': total_suggestions
                },
                'today': {
                    'votes': votes_today
                },
                'recent_activity': {
                    'votes_last_7_days': recent_votes,
                    'feedback_last_7_days': recent_feedback,
                    'complaints_last_7_days': recent_complaints
                },
                'popular_dishes': [
                    {'dish': dish, 'votes': count} 
                    for dish, count in popular_dishes
                ],
                'rating_distribution': [
                    {'rating': rating, 'count': count} 
                    for rating, count in rating_distribution
                ],
                'complaints_by_urgency': [
                    {'urgency': urgency, 'count': count} 
                    for urgency, count in complaints_by_urgency
                ]
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': {
                'code': 'INTERNAL_ERROR',
                'message': 'An error occurred while fetching dashboard stats'
            }
        }), 500

@admin_bp.route('/complaints/<int:complaint_id>/status', methods=['PUT'])
@require_admin_auth
def update_complaint_status():
    """Update complaint status"""
    try:
        complaint_id = request.view_args['complaint_id']
        data = request.json
        
        if not data or 'status' not in data:
            return jsonify({
                'success': False,
                'error': {
                    'code': 'MISSING_STATUS',
                    'message': 'Status is required'
                }
            }), 400
        
        valid_statuses = ['pending', 'in_progress', 'resolved', 'closed']
        if data['status'].lower() not in valid_statuses:
            return jsonify({
                'success': False,
                'error': {
                    'code': 'INVALID_STATUS',
                    'message': f'Status must be one of: {", ".join(valid_statuses)}'
                }
            }), 400
        
        complaint = Complaint.query.get_or_404(complaint_id)
        complaint.status = data['status'].lower()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Complaint status updated successfully',
            'complaint': complaint.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': {
                'code': 'INTERNAL_ERROR',
                'message': 'An error occurred while updating complaint status'
            }
        }), 500


# Excel export functionality
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    import pandas as pd
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: Excel export functionality not available (openpyxl/pandas not installed)")

from flask import send_file
import io

@admin_bp.route('/export/excel', methods=['GET'])
@require_admin_auth
def export_to_excel():
    """Export data to Excel format"""
    if not EXCEL_AVAILABLE:
        return jsonify({
            'success': False,
            'error': {
                'code': 'EXCEL_NOT_AVAILABLE',
                'message': 'Excel export functionality is not available on this deployment'
            }
        }), 503
    
    try:
        # Get query parameters
        export_type = request.args.get('type', 'all').lower()
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        def apply_header_style(ws, row_num=1):
            """Apply header styling to a worksheet"""
            for cell in ws[row_num]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        def add_data_to_sheet(ws, data, headers):
            """Add data to worksheet with headers"""
            # Add headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add data
            for row_idx, item in enumerate(data, 2):
                for col_idx, value in enumerate(item, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply header styling
            apply_header_style(ws)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Build date filters
        date_filter = {}
        if start_date:
            try:
                start_date_obj = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                date_filter['start'] = start_date_obj
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': {
                        'code': 'INVALID_DATE',
                        'message': 'Invalid start_date format. Use ISO format.'
                    }
                }), 400
        
        if end_date:
            try:
                end_date_obj = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                date_filter['end'] = end_date_obj
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': {
                        'code': 'INVALID_DATE',
                        'message': 'Invalid end_date format. Use ISO format.'
                    }
                }), 400
        
        # Export votes
        if export_type in ['all', 'votes']:
            query = Vote.query
            if 'start' in date_filter:
                query = query.filter(Vote.timestamp >= date_filter['start'])
            if 'end' in date_filter:
                query = query.filter(Vote.timestamp <= date_filter['end'])
            
            votes = query.order_by(Vote.timestamp.desc()).all()
            
            if votes:
                ws_votes = wb.create_sheet(title="Votes")
                vote_headers = ['ID', 'Day', 'Meal', 'Dish', 'IP Address', 'Timestamp', 'User Identifier']
                vote_data = [
                    [v.id, v.day, v.meal, v.dish, v.ip_address, 
                     v.timestamp.strftime('%Y-%m-%d %H:%M:%S') if v.timestamp else '', 
                     v.user_identifier[:16] + '...' if len(v.user_identifier) > 16 else v.user_identifier]
                    for v in votes
                ]
                add_data_to_sheet(ws_votes, vote_data, vote_headers)
        
        # Export feedback
        if export_type in ['all', 'feedback']:
            query = Feedback.query
            if 'start' in date_filter:
                query = query.filter(Feedback.timestamp >= date_filter['start'])
            if 'end' in date_filter:
                query = query.filter(Feedback.timestamp <= date_filter['end'])
            
            feedback_items = query.order_by(Feedback.timestamp.desc()).all()
            
            if feedback_items:
                ws_feedback = wb.create_sheet(title="Feedback")
                feedback_headers = ['ID', 'Type', 'Message', 'Rating', 'IP Address', 'Timestamp']
                feedback_data = [
                    [f.id, f.feedback_type, f.message, f.rating or 'N/A', f.ip_address,
                     f.timestamp.strftime('%Y-%m-%d %H:%M:%S') if f.timestamp else '']
                    for f in feedback_items
                ]
                add_data_to_sheet(ws_feedback, feedback_data, feedback_headers)
        
        # Export complaints
        if export_type in ['all', 'complaints']:
            query = Complaint.query
            if 'start' in date_filter:
                query = query.filter(Complaint.timestamp >= date_filter['start'])
            if 'end' in date_filter:
                query = query.filter(Complaint.timestamp <= date_filter['end'])
            
            complaints = query.order_by(Complaint.timestamp.desc()).all()
            
            if complaints:
                ws_complaints = wb.create_sheet(title="Complaints")
                complaint_headers = ['ID', 'Category', 'Message', 'Urgency', 'Status', 'IP Address', 'Timestamp']
                complaint_data = [
                    [c.id, c.category, c.message, c.urgency, c.status, c.ip_address,
                     c.timestamp.strftime('%Y-%m-%d %H:%M:%S') if c.timestamp else '']
                    for c in complaints
                ]
                add_data_to_sheet(ws_complaints, complaint_data, complaint_headers)
        
        # Export menu suggestions
        if export_type in ['all', 'suggestions', 'menu_suggestions']:
            query = MenuSuggestion.query
            if 'start' in date_filter:
                query = query.filter(MenuSuggestion.timestamp >= date_filter['start'])
            if 'end' in date_filter:
                query = query.filter(MenuSuggestion.timestamp <= date_filter['end'])
            
            suggestions = query.order_by(MenuSuggestion.timestamp.desc()).all()
            
            if suggestions:
                ws_suggestions = wb.create_sheet(title="Menu Suggestions")
                suggestion_headers = ['ID', 'Dish Name', 'Meal Type', 'Ingredients', 'Description', 'IP Address', 'Timestamp']
                suggestion_data = [
                    [s.id, s.dish_name, s.meal_type, s.ingredients or '', s.description or '', s.ip_address,
                     s.timestamp.strftime('%Y-%m-%d %H:%M:%S') if s.timestamp else '']
                    for s in suggestions
                ]
                add_data_to_sheet(ws_suggestions, suggestion_data, suggestion_headers)
        
        # Add summary sheet
        ws_summary = wb.create_sheet(title="Summary", index=0)
        
        # Summary data
        total_votes = Vote.query.count()
        total_feedback = Feedback.query.count()
        total_complaints = Complaint.query.count()
        total_suggestions = MenuSuggestion.query.count()
        
        # Popular dishes
        popular_dishes = db.session.query(
            Vote.dish,
            db.func.count(Vote.dish).label('vote_count')
        ).group_by(Vote.dish).order_by(
            db.func.count(Vote.dish).desc()
        ).limit(10).all()
        
        # Add summary information
        summary_data = [
            ['Metric', 'Value'],
            ['Total Votes', total_votes],
            ['Total Feedback', total_feedback],
            ['Total Complaints', total_complaints],
            ['Total Menu Suggestions', total_suggestions],
            ['Export Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')],
            ['Export Type', export_type.title()],
            ['', ''],
            ['Top 10 Popular Dishes', 'Vote Count']
        ]
        
        # Add popular dishes
        for dish, count in popular_dishes:
            summary_data.append([dish, count])
        
        # Add summary data to sheet
        for row_idx, (label, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=label)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        # Style the summary sheet
        apply_header_style(ws_summary, 1)
        apply_header_style(ws_summary, 9)  # Popular dishes header
        
        # Auto-adjust column widths for summary
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        # Check if workbook has any data sheets
        if len(wb.sheetnames) == 1:  # Only summary sheet
            return jsonify({
                'success': False,
                'error': {
                    'code': 'NO_DATA',
                    'message': 'No data available for the specified criteria'
                }
            }), 404
        
        # Save workbook to memory
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Generate filename
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f'mess_portal_data_{export_type}_{timestamp}.xlsx'
        
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': {
                'code': 'EXPORT_ERROR',
                'message': f'An error occurred while generating Excel file: {str(e)}'
            }
        }), 500

@admin_bp.route('/export/csv', methods=['GET'])
@require_admin_auth
def export_to_csv():
    """Export data to CSV format"""
    try:
        # Get query parameters
        export_type = request.args.get('type', 'votes').lower()
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        # Build date filters
        date_filter = {}
        if start_date:
            try:
                start_date_obj = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
                date_filter['start'] = start_date_obj
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': {
                        'code': 'INVALID_DATE',
                        'message': 'Invalid start_date format. Use ISO format.'
                    }
                }), 400
        
        if end_date:
            try:
                end_date_obj = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
                date_filter['end'] = end_date_obj
            except ValueError:
                return jsonify({
                    'success': False,
                    'error': {
                        'code': 'INVALID_DATE',
                        'message': 'Invalid end_date format. Use ISO format.'
                    }
                }), 400
        
        # Get data based on export type
        if export_type == 'votes':
            query = Vote.query
            if 'start' in date_filter:
                query = query.filter(Vote.timestamp >= date_filter['start'])
            if 'end' in date_filter:
                query = query.filter(Vote.timestamp <= date_filter['end'])
            
            data = query.order_by(Vote.timestamp.desc()).all()
            csv_data = [vote.to_dict() for vote in data]
            
        elif export_type == 'feedback':
            query = Feedback.query
            if 'start' in date_filter:
                query = query.filter(Feedback.timestamp >= date_filter['start'])
            if 'end' in date_filter:
                query = query.filter(Feedback.timestamp <= date_filter['end'])
            
            data = query.order_by(Feedback.timestamp.desc()).all()
            csv_data = [feedback.to_dict() for feedback in data]
            
        elif export_type == 'complaints':
            query = Complaint.query
            if 'start' in date_filter:
                query = query.filter(Complaint.timestamp >= date_filter['start'])
            if 'end' in date_filter:
                query = query.filter(Complaint.timestamp <= date_filter['end'])
            
            data = query.order_by(Complaint.timestamp.desc()).all()
            csv_data = [complaint.to_dict() for complaint in data]
            
        elif export_type == 'menu_suggestions':
            query = MenuSuggestion.query
            if 'start' in date_filter:
                query = query.filter(MenuSuggestion.timestamp >= date_filter['start'])
            if 'end' in date_filter:
                query = query.filter(MenuSuggestion.timestamp <= date_filter['end'])
            
            data = query.order_by(MenuSuggestion.timestamp.desc()).all()
            csv_data = [suggestion.to_dict() for suggestion in data]
            
        else:
            return jsonify({
                'success': False,
                'error': {
                    'code': 'INVALID_TYPE',
                    'message': 'Invalid export type. Use: votes, feedback, complaints, or menu_suggestions'
                }
            }), 400
        
        if not csv_data:
            return jsonify({
                'success': False,
                'error': {
                    'code': 'NO_DATA',
                    'message': 'No data available for the specified criteria'
                }
            }), 404
        
        # Convert to DataFrame and then to CSV
        df = pd.DataFrame(csv_data)
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False)
        csv_buffer.seek(0)
        
        # Generate filename
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f'mess_portal_{export_type}_{timestamp}.csv'
        
        # Convert StringIO to BytesIO for send_file
        csv_bytes = io.BytesIO()
        csv_bytes.write(csv_buffer.getvalue().encode('utf-8'))
        csv_bytes.seek(0)
        
        return send_file(
            csv_bytes,
            as_attachment=True,
            download_name=filename,
            mimetype='text/csv'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': {
                'code': 'EXPORT_ERROR',
                'message': f'An error occurred while generating CSV file: {str(e)}'
            }
        }), 500

